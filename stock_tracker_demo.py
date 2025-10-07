#!/usr/bin/env python3
"""
KarthiStocks Demo - Live Stock Price Tracker with Mock Data
This version uses mock data to demonstrate functionality without internet access
"""

import openpyxl
import pandas as pd
from datetime import datetime
import sys
import os
import random


def read_stock_symbols(excel_file, sheet_name='Stocks', symbol_column='A'):
    """Read stock symbols from Excel file"""
    try:
        workbook = openpyxl.load_workbook(excel_file)
        if sheet_name not in workbook.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found. Using first sheet: {workbook.sheetnames[0]}")
            sheet_name = workbook.sheetnames[0]
        
        sheet = workbook[sheet_name]
        symbols = []
        
        # Read symbols from column (skip header in row 1)
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet[f'{symbol_column}{row}'].value
            if cell_value and str(cell_value).strip() != 'Error':
                symbols.append(str(cell_value).strip())
        
        workbook.close()
        return symbols
    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file}' not found.")
        sys.exit(1)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        sys.exit(1)


def fetch_live_prices_mock(symbols):
    """Fetch mock stock prices for demonstration"""
    # Mock data for common stocks
    mock_prices = {
        'AAPL': {'name': 'Apple Inc.', 'base_price': 175.43, 'prev': 174.25, 'volume': 45234567, 'cap': 2750000000000},
        'GOOGL': {'name': 'Alphabet Inc.', 'base_price': 139.82, 'prev': 138.95, 'volume': 23456789, 'cap': 1750000000000},
        'MSFT': {'name': 'Microsoft Corporation', 'base_price': 338.11, 'prev': 337.45, 'volume': 34567890, 'cap': 2510000000000},
        'AMZN': {'name': 'Amazon.com Inc.', 'base_price': 145.28, 'prev': 144.82, 'volume': 28901234, 'cap': 1500000000000},
        'TSLA': {'name': 'Tesla Inc.', 'base_price': 242.15, 'prev': 245.30, 'volume': 67890123, 'cap': 765000000000},
    }
    
    stock_data = []
    
    print(f"\nFetching live prices for {len(symbols)} stocks...\n")
    
    for symbol in symbols:
        try:
            # Use mock data if available, otherwise generate random data
            if symbol in mock_prices:
                data = mock_prices[symbol]
                # Add small random variation to simulate live prices
                variation = random.uniform(-0.02, 0.02)
                current_price = data['base_price'] * (1 + variation)
                previous_close = data['prev']
                company = data['name']
                volume = data['volume']
                market_cap = data['cap']
            else:
                # Generate random data for unknown symbols
                company = f"{symbol} Corporation"
                current_price = random.uniform(50, 500)
                previous_close = current_price * random.uniform(0.98, 1.02)
                volume = random.randint(1000000, 100000000)
                market_cap = random.randint(1000000000, 3000000000000)
            
            change = current_price - previous_close
            change_percent = (change / previous_close) * 100
            
            stock_data.append({
                'Symbol': symbol,
                'Company': company,
                'Current Price': round(current_price, 2),
                'Previous Close': round(previous_close, 2),
                'Change': round(change, 2),
                'Change %': round(change_percent, 2),
                'Volume': volume,
                'Market Cap': market_cap
            })
        except Exception as e:
            print(f"Error processing {symbol}: {e}")
    
    return stock_data


def update_excel(excel_file, stock_data, sheet_name='Stocks'):
    """Update Excel file with live stock prices"""
    try:
        workbook = openpyxl.load_workbook(excel_file)
        if sheet_name not in workbook.sheetnames:
            sheet_name = workbook.sheetnames[0]
        
        sheet = workbook[sheet_name]
        
        # Update headers
        headers = ['Symbol', 'Company', 'Current Price', 'Previous Close', 'Change', 'Change %', 'Volume', 'Market Cap', 'Last Updated']
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        
        # Update data
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for row, data in enumerate(stock_data, start=2):
            sheet.cell(row=row, column=1, value=data['Symbol'])
            sheet.cell(row=row, column=2, value=data['Company'])
            sheet.cell(row=row, column=3, value=data['Current Price'])
            sheet.cell(row=row, column=4, value=data['Previous Close'])
            sheet.cell(row=row, column=5, value=data['Change'])
            sheet.cell(row=row, column=6, value=data['Change %'])
            sheet.cell(row=row, column=7, value=data['Volume'])
            sheet.cell(row=row, column=8, value=data['Market Cap'])
            sheet.cell(row=row, column=9, value=timestamp)
        
        workbook.save(excel_file)
        workbook.close()
        print(f"\n✓ Excel file updated successfully: {excel_file}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")


def display_stock_prices(stock_data):
    """Display stock prices in a formatted table"""
    df = pd.DataFrame(stock_data)
    
    # Format the display
    pd.options.display.float_format = '{:,.2f}'.format
    
    print("\n" + "="*120)
    print(" "*45 + "LIVE STOCK PRICES")
    print("="*120)
    print(df.to_string(index=False))
    print("="*120)
    print(f"\n✓ Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")


def main():
    """Main function"""
    excel_file = 'stocks.xlsx'
    
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    
    if not os.path.exists(excel_file):
        print(f"Error: Excel file '{excel_file}' not found.")
        print(f"\nUsage: python stock_tracker_demo.py [excel_file_path]")
        sys.exit(1)
    
    print("="*120)
    print(" "*38 + "KarthiStocks - Live Stock Price Tracker (Demo)")
    print("="*120)
    print("\nNOTE: This demo uses mock data to demonstrate functionality")
    
    # Read stock symbols from Excel
    symbols = read_stock_symbols(excel_file)
    
    if not symbols:
        print("No stock symbols found in Excel file.")
        sys.exit(1)
    
    print(f"\n✓ Found {len(symbols)} stock symbols: {', '.join(symbols)}")
    
    # Fetch mock prices
    stock_data = fetch_live_prices_mock(symbols)
    
    # Display prices
    display_stock_prices(stock_data)
    
    # Update Excel file
    update_excel(excel_file, stock_data)
    
    print("✓ Done!\n")


if __name__ == '__main__':
    main()
