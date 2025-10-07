#!/usr/bin/env python3
"""
KarthiStocks - Live Stock Price Tracker
Connects to Excel worksheet, fetches live stock prices, and displays them
"""

import openpyxl
import yfinance as yf
import pandas as pd
from datetime import datetime
import sys
import os


def read_stock_symbols(excel_file, sheet_name='Stocks', symbol_column='A'):
    """
    Read stock symbols from Excel file
    
    Args:
        excel_file: Path to Excel file
        sheet_name: Name of the worksheet (default: 'Stocks')
        symbol_column: Column containing stock symbols (default: 'A')
    
    Returns:
        List of stock symbols
    """
    try:
        workbook = openpyxl.load_workbook(excel_file)
        if sheet_name not in workbook.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found. Using first sheet: {workbook.sheetnames[0]}")
            sheet_name = workbook.sheetnames[0]
        
        sheet = workbook[sheet_name]
        symbols = []
        
        # Read symbols from column (skip header in row 1)
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet[f'{symbol_column}{row}'].value
            if cell_value:
                symbols.append(str(cell_value).strip())
        
        workbook.close()
        return symbols
    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file}' not found.")
        sys.exit(1)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        sys.exit(1)


def fetch_live_prices(symbols):
    """
    Fetch live stock prices using yfinance
    
    Args:
        symbols: List of stock symbols
    
    Returns:
        Dictionary with stock data
    """
    stock_data = []
    
    print(f"\nFetching live prices for {len(symbols)} stocks...\n")
    
    for symbol in symbols:
        try:
            ticker = yf.Ticker(symbol)
            info = ticker.info
            history = ticker.history(period='1d')
            
            if not history.empty:
                current_price = history['Close'].iloc[-1]
                previous_close = info.get('previousClose', current_price)
                change = current_price - previous_close
                change_percent = (change / previous_close) * 100 if previous_close else 0
                
                stock_data.append({
                    'Symbol': symbol,
                    'Company': info.get('shortName', symbol),
                    'Current Price': round(current_price, 2),
                    'Previous Close': round(previous_close, 2),
                    'Change': round(change, 2),
                    'Change %': round(change_percent, 2),
                    'Volume': info.get('volume', 0),
                    'Market Cap': info.get('marketCap', 0)
                })
            else:
                print(f"Warning: No data available for {symbol}")
                stock_data.append({
                    'Symbol': symbol,
                    'Company': 'N/A',
                    'Current Price': 'N/A',
                    'Previous Close': 'N/A',
                    'Change': 'N/A',
                    'Change %': 'N/A',
                    'Volume': 'N/A',
                    'Market Cap': 'N/A'
                })
        except Exception as e:
            print(f"Error fetching data for {symbol}: {e}")
            stock_data.append({
                'Symbol': symbol,
                'Company': 'Error',
                'Current Price': 'N/A',
                'Previous Close': 'N/A',
                'Change': 'N/A',
                'Change %': 'N/A',
                'Volume': 'N/A',
                'Market Cap': 'N/A'
            })
    
    return stock_data


def update_excel(excel_file, stock_data, sheet_name='Stocks'):
    """
    Update Excel file with live stock prices
    
    Args:
        excel_file: Path to Excel file
        stock_data: List of dictionaries with stock data
        sheet_name: Name of the worksheet (default: 'Stocks')
    """
    try:
        workbook = openpyxl.load_workbook(excel_file)
        if sheet_name not in workbook.sheetnames:
            sheet_name = workbook.sheetnames[0]
        
        sheet = workbook[sheet_name]
        
        # Update headers if not present
        headers = ['Symbol', 'Company', 'Current Price', 'Previous Close', 'Change', 'Change %', 'Volume', 'Market Cap', 'Last Updated']
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        
        # Update data
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for row, data in enumerate(stock_data, start=2):
            sheet.cell(row=row, column=1, value=data['Symbol'])
            sheet.cell(row=row, column=2, value=data['Company'])
            sheet.cell(row=row, column=3, value=data['Current Price'])
            sheet.cell(row=row, column=4, value=data['Previous Close'])
            sheet.cell(row=row, column=5, value=data['Change'])
            sheet.cell(row=row, column=6, value=data['Change %'])
            sheet.cell(row=row, column=7, value=data['Volume'])
            sheet.cell(row=row, column=8, value=data['Market Cap'])
            sheet.cell(row=row, column=9, value=timestamp)
        
        workbook.save(excel_file)
        workbook.close()
        print(f"\nExcel file updated successfully: {excel_file}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")


def display_stock_prices(stock_data):
    """
    Display stock prices in a formatted table
    
    Args:
        stock_data: List of dictionaries with stock data
    """
    df = pd.DataFrame(stock_data)
    
    print("\n" + "="*100)
    print("LIVE STOCK PRICES")
    print("="*100)
    print(df.to_string(index=False))
    print("="*100)
    print(f"\nLast Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")


def main():
    """Main function"""
    # Default Excel file path
    excel_file = 'stocks.xlsx'
    
    # Check if custom file path provided
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    
    if not os.path.exists(excel_file):
        print(f"Error: Excel file '{excel_file}' not found.")
        print(f"\nUsage: python stock_tracker.py [excel_file_path]")
        print(f"Example: python stock_tracker.py stocks.xlsx")
        sys.exit(1)
    
    print("="*100)
    print("KarthiStocks - Live Stock Price Tracker")
    print("="*100)
    
    # Read stock symbols from Excel
    symbols = read_stock_symbols(excel_file)
    
    if not symbols:
        print("No stock symbols found in Excel file.")
        sys.exit(1)
    
    print(f"Found {len(symbols)} stock symbols: {', '.join(symbols)}")
    
    # Fetch live prices
    stock_data = fetch_live_prices(symbols)
    
    # Display prices
    display_stock_prices(stock_data)
    
    # Update Excel file
    update_excel(excel_file, stock_data)
    
    print("Done!")


if __name__ == '__main__':
    main()
